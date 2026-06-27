"""
Import claims triangles (and vectors such as earned premium) from Excel into a
single tidy "long" table, and convert that table back into the visual triangle
arrays the reserving tools consume.

Storage format (long / tidy) - one row per observed cell:

    Region | Entity | LoB | AY | DY | Type | Name | Value

  * AY   accident year (or any origin label)
  * DY   development length / lag (1, 2, 3, ...); <NA> for vectors
  * Type "Triangle" or "Vector"
  * Name identifier for the triangle/vector
  * Value INCREMENTAL value (cumulative sources are converted on import)

One table can hold many triangles and their premium/ultimate vectors. To run a
tool, filter to one triangle and pivot to a 2D array with to_triangle_array().

The importer is driven by a list of "specs", each describing where one triangle
lives and its metadata. Triangles may be in different files, sheets, ranges and
sizes. Layout assumed is a labelled grid (an AY label column down the side and a
DY header row across the top); set has_labels=False for a raw value block.

Spec keys (a `defaults` dict can supply any of these for every spec):
    file        path to .xlsx                                   (required)
    sheet       worksheet name                                  (required)
    name        triangle/vector identifier                      (required)
    region, entity, lob   metadata (default "")
    type        "Triangle" (default) or "Vector"
    range       A1 range incl. labels, e.g. "B2:L12"   } give one
    top_left    corner cell, e.g. "B2", auto-detect extent }
    has_labels  default True (labelled grid)
    orientation "incremental" (default) or "cumulative"
    dev_as      "length" (default; DY = 1..n) or "label" (use the DY header row)

Example:
    from triangle_io import import_triangles, to_triangle
    long = import_triangles(
        [
            {"file": "2024.xlsx", "sheet": "GL",  "range": "B2:L12",
             "region": "Canada", "lob": "GL", "name": "GL_2024"},
            {"file": "2024.xlsx", "sheet": "Auto", "top_left": "C3",
             "region": "Canada", "lob": "Auto", "name": "Auto_2024",
             "orientation": "cumulative"},
        ],
        defaults={"entity": "ABC"},
    )
    tri = to_triangle(long, name="GL_2024")   # labelled incremental triangle
    # pass straight to a tool: merz_wuthrich(tri) / risk_emergence(tri)
"""

from __future__ import annotations

import argparse

import numpy as np
import pandas as pd

LONG_COLUMNS = ["Region", "Entity", "LoB", "AY", "DY", "Type", "Name", "Value"]


# ---------------------------------------------------------------------------
# Low-level Excel reading
# ---------------------------------------------------------------------------

def _to_float(v):
    """Excel cell value -> float, with blanks/None/strings -> NaN."""
    if v is None or v == "":
        return np.nan
    try:
        return float(v)
    except (TypeError, ValueError):
        return np.nan


def _read_block(ws, spec):
    """
    Return (dy_labels, ay_labels, value_matrix) for one triangle/vector.

    Uses an explicit `range` if given, else auto-detects the extent from a
    `top_left` corner by scanning the (fully populated) DY header row and AY
    label column.
    """
    has_labels = spec.get("has_labels", True)

    if spec.get("range"):
        cells = ws[spec["range"]]
        grid = [[c.value for c in row] for row in cells]
    elif spec.get("top_left"):
        from openpyxl.utils import coordinate_to_tuple
        r0, c0 = coordinate_to_tuple(spec["top_left"])
        if not has_labels:
            raise ValueError("top_left auto-detect requires has_labels=True; "
                             "use an explicit range for raw value blocks.")
        ncols = 0
        while ws.cell(r0, c0 + 1 + ncols).value not in (None, ""):
            ncols += 1
        nrows = 0
        while ws.cell(r0 + 1 + nrows, c0).value not in (None, ""):
            nrows += 1
        grid = [[ws.cell(r0 + i, c0 + j).value for j in range(ncols + 1)]
                for i in range(nrows + 1)]
    else:
        raise ValueError(f"Spec for '{spec.get('name')}' needs 'range' or 'top_left'.")

    if has_labels:
        dy_labels = grid[0][1:]
        ay_labels = [row[0] for row in grid[1:]]
        values = [row[1:] for row in grid[1:]]
    else:
        dy_labels = list(range(1, len(grid[0]) + 1))
        ay_labels = list(range(1, len(grid) + 1))
        values = grid
    return dy_labels, ay_labels, values


def _coerce_label(x):
    """Keep integer-looking labels as ints (e.g. 2015), else leave as-is."""
    f = _to_float(x)
    if not np.isnan(f) and float(f).is_integer():
        return int(f)
    return x


def _cumulative_to_incremental(row_vals):
    """Difference a cumulative row over its observed (contiguous) prefix."""
    out = [np.nan] * len(row_vals)
    prev = 0.0
    for j, v in enumerate(row_vals):
        if np.isnan(v):
            break
        out[j] = v - prev
        prev = v
    return out


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def import_triangles(specs, defaults: dict | None = None) -> pd.DataFrame:
    """Read one or more triangles/vectors from Excel into the long table."""
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise ImportError("openpyxl is required to read Excel files "
                          "(`pip install openpyxl`).") from e

    if isinstance(specs, dict):
        specs = [specs]
    defaults = defaults or {}

    workbooks: dict[str, object] = {}
    records = []

    for raw in specs:
        s = {**defaults, **raw}
        for req in ("file", "sheet", "name"):
            if req not in s:
                raise ValueError(f"Spec missing required key '{req}': {raw}")

        if s["file"] not in workbooks:
            workbooks[s["file"]] = load_workbook(s["file"], data_only=True, read_only=True)
        ws = workbooks[s["file"]][s["sheet"]]

        dy_labels, ay_labels, values = _read_block(ws, s)
        ttype = s.get("type", "Triangle")
        meta = {"Region": s.get("region", ""), "Entity": s.get("entity", ""),
                "LoB": s.get("lob", ""), "Type": ttype, "Name": s["name"]}

        if ttype == "Vector":
            for ay, row in zip(ay_labels, values):
                v = _to_float(row[0] if isinstance(row, (list, tuple)) else row)
                if np.isnan(v):
                    continue
                records.append({**meta, "AY": _coerce_label(ay), "DY": pd.NA, "Value": v})
            continue

        # Triangle
        matrix = [[_to_float(v) for v in row] for row in values]
        if s.get("orientation", "incremental") == "cumulative":
            matrix = [_cumulative_to_incremental(r) for r in matrix]

        dev_as = s.get("dev_as", "length")
        for ay, row in zip(ay_labels, matrix):
            for j, v in enumerate(row):
                if np.isnan(v):
                    continue
                dy = _coerce_label(dy_labels[j]) if dev_as == "label" else j + 1
                records.append({**meta, "AY": _coerce_label(ay), "DY": dy, "Value": v})

    df = pd.DataFrame(records, columns=LONG_COLUMNS)
    if not df.empty:
        df["AY"] = pd.array(df["AY"], dtype="Int64") if df["AY"].map(
            lambda x: isinstance(x, int)).all() else df["AY"]
        df["DY"] = pd.array(df["DY"], dtype="Int64")
    return df


def to_triangle(long_df: pd.DataFrame, region=None, entity=None, lob=None,
                name=None) -> pd.DataFrame:
    """
    Filter the long table to a single triangle and pivot it to an incremental
    triangle, returned as a labelled DataFrame (AY index, DY columns; missing
    cells are NaN).

    The result can be passed straight to merz_wuthrich() / risk_emergence(),
    which accept a DataFrame or array. Use .to_numpy() if you want the raw array,
    or .index / .columns for the AY / DY labels.
    """
    sub = long_df[long_df["Type"] == "Triangle"]
    for col, val in [("Region", region), ("Entity", entity), ("LoB", lob), ("Name", name)]:
        if val is not None:
            sub = sub[sub[col] == val]

    names = sub["Name"].unique()
    if len(names) == 0:
        raise ValueError("No matching triangle found for the given filters.")
    if len(names) > 1:
        raise ValueError(f"Filters match multiple triangles {list(names)}; "
                         "narrow region/entity/lob/name.")

    pivot = sub.pivot_table(index="AY", columns="DY", values="Value", aggfunc="sum")
    return pivot.sort_index().sort_index(axis=1)


# ---------------------------------------------------------------------------
# CLI: read a manifest (YAML or JSON) of specs -> write the combined long CSV
# ---------------------------------------------------------------------------

def _load_manifest(path: str) -> tuple[list, dict]:
    if path.lower().endswith((".yaml", ".yml")):
        import yaml
        with open(path) as f:
            data = yaml.safe_load(f)
    else:
        import json
        with open(path) as f:
            data = json.load(f)
    return data.get("triangles", []), data.get("defaults", {})


def main():
    parser = argparse.ArgumentParser(description="Import Excel triangles into a long table.")
    parser.add_argument("--manifest", required=True,
                        help="YAML/JSON manifest with 'triangles' (list of specs) and optional 'defaults'.")
    parser.add_argument("--out", required=True, help="Output long-format CSV path.")
    args = parser.parse_args()

    specs, defaults = _load_manifest(args.manifest)
    long = import_triangles(specs, defaults=defaults)
    long.to_csv(args.out, index=False)
    n_tri = long[long["Type"] == "Triangle"]["Name"].nunique()
    n_vec = long[long["Type"] == "Vector"]["Name"].nunique()
    print(f"Imported {n_tri} triangle(s) and {n_vec} vector(s), "
          f"{len(long)} rows -> {args.out}")


if __name__ == "__main__":
    main()
